"""Excel 投递跟踪表生成与更新"""

import json
from pathlib import Path
from core import PROJECT_ROOT, DATA_DIR

EXCEL_PATH = DATA_DIR / "秋招投递跟踪表.xlsx"


def generate():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "投递跟踪表"

    hfont = Font(bold=True, size=11, color='FFFFFF')
    hfill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['公司', '梯队', '类别', '部门/团队', '岗位名称', '工作地点',
               'JD核心要求', '匹配度', '投递链接', '内推信息', '网申开放',
               '网申截止', '薪资范围', '投递状态', '笔试时间', '面试进度', '备注']
    widths = [12, 6, 14, 16, 22, 10, 30, 8, 35, 20, 12, 12, 12, 10, 12, 15, 25]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font, cell.fill, cell.border = hfont, hfill, border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    company_path = PROJECT_ROOT / "公司清单.json"
    if company_path.exists():
        with open(company_path, encoding="utf-8") as f:
            companies = json.load(f).get("companies", [])
        for row, c in enumerate(companies, 2):
            ws.cell(row=row, column=1, value=c["name"]).border = border
            ws.cell(row=row, column=2, value=c.get("tier", "")).border = border
            ws.cell(row=row, column=3, value=c.get("category", "")).border = border
            ws.cell(row=row, column=14, value="待关注").border = border

    ws.freeze_panes = 'A2'
    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_PATH)
    print(f"[OK] 投递跟踪表已生成: {EXCEL_PATH}")
